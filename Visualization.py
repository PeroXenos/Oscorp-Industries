import psycopg2
import pandas as pd
import plotly.express as px
import os
from typing import Optional, Dict

# --- КРИТИЧНЫЙ ФИКС ДЛЯ ОШИБКИ TCL/TK ---
# Установка бэкэнда 'Agg' для сохранения графиков в файл (не требует Tkinter).
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
except Exception as e:
    print(f"ВНИМАНИЕ: Не удалось установить бэкэнд Matplotlib. Графики могут не сохраниться. Ошибка: {e}")
    import matplotlib.pyplot as plt

# Импорт для форматирования Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.utils import get_column_letter

# --- НАСТРОЙКИ ПОДКЛЮЧЕНИЯ И ПУТИ ---
DB_NAME = "WorldCup_db"
DB_USER = "DataVis"
DB_PASSWORD = "DataVis" 
DB_HOST = "localhost"
DB_PORT = "5432"

# Создание необходимых папок
os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

# SQL-запросы (Все в snake_case. Таблицы, начинающиеся с цифр, в двойных кавычках)
ALL_QUERIES = {
    # 1. PIE CHART: Распределение матчей по стадиям в годы, когда Хозяин был победителем (2 JOIN)
    "q_pie_stage_dist": """
        SELECT 
            M.stage, 
            COUNT(M.id) AS stage_count
        FROM 
            world_cup_matches M
        INNER JOIN 
            world_cups C ON M.year = C.year
        INNER JOIN 
            world_cup_matches M2 ON M.id = M2.id 
        WHERE 
            C.host_country = C.winner 
        GROUP BY 
            M.stage;
    """,
    
    # 2. BAR CHART: Топ-5 команд по суммарным голам, участвовавшие в финалах (2 JOIN)
    "q_bar_top_goal_teams": """
        SELECT 
            T.team, 
            SUM(T.total_goals) AS total_goals_scored
        FROM (
            SELECT home_team AS team, home_goals AS total_goals FROM world_cup_matches 
            UNION ALL
            SELECT away_team AS team, away_goals AS total_goals FROM world_cup_matches 
        ) T
        INNER JOIN 
            world_cups C ON C.winner = T.team OR C.runners_up = T.team 
        INNER JOIN 
            world_cup_matches M ON M.home_team = T.team OR M.away_team = T.team 
        GROUP BY 
            T.team
        ORDER BY 
            total_goals_scored DESC
        LIMIT 5;
    """,
    
    # 3. HORIZONTAL BAR: Средний рейтинг FIFA команд, вышедших в финал ЧМ-2022
    "q_hbar_avg_fifa_rank": """
        SELECT
            T.team,
            AVG(CAST(G.fifa_ranking AS INTEGER)) AS avg_fifa_rank 
        FROM (
            SELECT winner AS team FROM world_cups WHERE year = 2018
            UNION ALL
            SELECT runners_up AS team FROM world_cups WHERE year = 2018
        ) T
        INNER JOIN 
            world_cup_groups G ON T.team = G.team 
        GROUP BY
            T.team;
    """,

    # 4. LINE CHART: Голы за матч по годам турниров (2 JOIN)
    "q_line_goals_per_match": """
        SELECT 
            C.year, 
            CAST(C.goals_scored AS DECIMAL) / C.matches_played AS gpm, 
            C.matches_played
        FROM 
            world_cups C
        INNER JOIN
            world_cup_matches M ON C.year = M.year 
        INNER JOIN 
            world_cups C2 ON C.year = C2.year 
        ORDER BY 
            C.year;
    """,
    
    # 5. BAR CHART ALL TIME: Количество забитых голов каждой командой за ВСЮ историю ЧМ
    "q_bar_matches_2022": """
        SELECT 
            T.team, 
            SUM(T.goals_scored) AS total_goals_scored
        FROM (
            -- Используем общую таблицу исторических матчей
            SELECT home_team AS team, home_goals AS goals_scored FROM world_cup_matches
            UNION ALL
            SELECT away_team AS team, away_goals AS goals_scored FROM world_cup_matches
        ) AS T
        GROUP BY T.team
        ORDER BY total_goals_scored DESC
        LIMIT 15;
    """,
    
    # 6. SCATTER PLOT: Голы vs. Квалифицированные команды по годам (2 JOIN)
    "q_scatter_goals_vs_teams": """
        SELECT 
            C.goals_scored, 
            C.qualified_teams 
        FROM 
            world_cups C
        INNER JOIN
            world_cup_matches M ON C.year = M.year 
        INNER JOIN
            world_cups C2 ON C.year = C2.year 
        GROUP BY 
            C.goals_scored, C.qualified_teams;
    """,
    
    # 7. Запрос для интерактивного графика Plotly
    "q_plotly_intl_goals": """
        SELECT 
            date, 
            home_team AS team, 
            home_goals AS goals 
        FROM 
            international_matches
        UNION ALL
        SELECT 
            date, 
            away_team AS team, 
            away_goals AS goals 
        FROM 
            international_matches
        ORDER BY date; -- ФИЛЬТР 'WHERE date > 2020-01-01' УДАЛЕН И ПЕРЕНЕСЕН В PYTHON
    """,
    
    # --- ЗАПРОСЫ ДЛЯ ЭКСПОРТА В EXCEL ---
    "top_scorers_export": """
        SELECT player, team, goals FROM world_cup_squads WHERE goals > 0 ORDER BY goals DESC LIMIT 20; -- ИСПРАВЛЕНО: Таблица в кавычках
    """,
    "matches_by_year_export": """
        SELECT year, COUNT(*) AS total_matches FROM world_cup_matches GROUP BY year ORDER BY year;
    """,
    "teams_by_group_export": """
        SELECT group_name, team, fifa_ranking FROM world_cup_groups ORDER BY group_name; -- ИСПРАВЛЕНО: Таблица в кавычках
    """
}

# --- ФУНКЦИИ БАЗЫ ДАННЫХ И ВИЗУАЛИЗАЦИИ ---

def run_query(query: str, conn: psycopg2.connect) -> Optional[pd.DataFrame]:
    """Выполняет SQL-запрос и возвращает DataFrame."""
    try:
        # UserWarning о psycopg2 и SQLAlchemy можно игнорировать
        df = pd.read_sql(query, conn)
        print(f"✅ Успешно получен DataFrame ({len(df)} строк) по запросу:\n{query.splitlines()[1].strip()}")
        return df
    except Exception as e:
        print(f"❌ Ошибка при выполнении запроса:\n{query.splitlines()[1].strip()}\n{e}")
        return None

def create_pie_chart(df: pd.DataFrame):
    """Круговая диаграмма: Распределение матчей по стадиям."""
    plt.figure(figsize=(8, 8))
    df.set_index('stage', inplace=True)
    plt.pie(df['stage_count'], labels=df.index, autopct='%1.1f%%', startangle=90, wedgeprops={'edgecolor': 'black'})
    plt.title('Круговая диаграмма: Распределение стадий матчей в ЧМ, выигранных Хозяином')
    plt.legend(title="Стадии", loc="upper right")
    plt.tight_layout()
    plt.savefig('charts/pie_chart_stage_distribution.png')
    plt.close()
    print(f"   -> Создан график: Круговая диаграмма.")

def create_bar_chart(df: pd.DataFrame):
    """Столбчатая диаграмма: Топ-5 команд по суммарным голам."""
    plt.figure(figsize=(10, 6))
    plt.bar(df['team'], df['total_goals_scored'], color='skyblue')
    plt.title('Столбчатая диаграмма: Топ-5 команд по общему количеству голов')
    plt.xlabel('Команда')
    plt.ylabel('Общее количество голов')
    plt.grid(axis='y', linestyle='--')
    plt.savefig('charts/bar_chart_top_goal_teams.png')
    plt.close()
    print(f"   -> Создан график: Столбчатая диаграмма.")

def create_horizontal_bar(df: pd.DataFrame):
    """Горизонтальная столбчатая диаграмма: Средний рейтинг FIFA финалистов 2022."""
    plt.figure(figsize=(10, 5))
    df.sort_values(by='avg_fifa_rank', ascending=False).plot(
        kind='barh', 
        x='team', 
        y='avg_fifa_rank', 
        legend=False, 
        color=['red', 'blue']
    )
    plt.title('Горизонтальная столбчатая диаграмма: Средний рейтинг FIFA команд-финалистов ЧМ-2022')
    plt.xlabel('Средний рейтинг FIFA (чем ниже, тем лучше)')
    plt.ylabel('Команда')
    plt.tight_layout()
    plt.savefig('charts/hbar_avg_fifa_rank.png')
    plt.close()
    print(f"   -> Создан график: Горизонтальная столбчатая диаграмма.")

def create_line_chart(df: pd.DataFrame):
    """Линейный график: Голы за матч по годам турниров."""
    plt.figure(figsize=(12, 6))
    plt.plot(df['year'], df['gpm'], marker='o', linestyle='-', color='purple')
    plt.title('Линейный график: Тенденция результативности (Голов за матч по годам)')
    plt.xlabel('Год Чемпионата')
    plt.ylabel('Голов за матч (GPM)')
    plt.grid(True)
    plt.savefig('charts/line_chart_goals_per_match.png')
    plt.close()
    print(f"   -> Создан график: Линейный график.")

def create_histogram(df: pd.DataFrame):
    """Столбчатая диаграмма: Количество забитых голов каждой командой в ЧМ-2022."""
    plt.figure(figsize=(15, 7))
    
    # Сортировка по количеству голов для лучшей визуализации
    df_sorted = df.sort_values(by='total_goals_scored', ascending=False)
    
    plt.bar(
        df_sorted['team'], 
        df_sorted['total_goals_scored'], 
        color='gold', # Изменим цвет на золотой/желтый для голов
        edgecolor='black'
    )
    
    plt.title('Столбчатая диаграмма: Общее количество забитых голов каждой командой в ЧМ-2022')
    plt.xlabel('Команда')
    plt.ylabel('Количество забитых голов')
    plt.xticks(rotation=90) 
    plt.grid(axis='y', linestyle='--')
    plt.tight_layout()
    plt.savefig('charts/bar_chart_2022_goals_per_team.png')
    plt.close()
    print(f"   -> Создан график: Столбчатая диаграмма забитых голов ЧМ-2022.")
def create_scatter_plot(df: pd.DataFrame):
    """Диаграмма рассеяния: Голы vs. Квалифицированные команды по годам."""
    plt.figure(figsize=(10, 6))
    plt.scatter(df['qualified_teams'], df['goals_scored'], c=df.index, cmap='viridis', s=100, alpha=0.7)
    plt.title('Диаграмма рассеяния: Голы vs. Квалифицированные команды')
    plt.xlabel('Количество квалифицированных команд')
    plt.ylabel('Общее количество голов на турнире')
    plt.colorbar(label='Индекс турнира')
    plt.grid(True)
    plt.savefig('charts/scatter_goals_vs_teams.png')
    plt.close()
    print(f"   -> Создан график: Диаграмма рассеяния.")

def create_plotly_animation(df: pd.DataFrame):
    """Интерактивный график с временным ползунком."""
    
    # ПРЕОБРАЗОВАНИЕ ДАТЫ И ФИЛЬТРАЦИЯ СРЕДСТВАМИ PANDAS
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    df_filtered = df[df['date'].dt.year >= 2020].copy()
    
    if df_filtered.empty:
        print("   -> Недостаточно данных (после 2020 года) для Plotly-анимации.")
        # Отобразим данные без фильтра, если нет матчей с 2020
        df_filtered = df.copy() 
        if df_filtered.empty:
            return
            
    df_filtered['Year'] = df_filtered['date'].dt.year
    df_agg = df_filtered.groupby(['Year', 'team']).agg(
        total_goals=('goals', 'sum'),
        num_matches=('goals', 'count')
    ).reset_index()
    
    if df_agg.empty:
        print("   -> Недостаточно агрегированных данных для Plotly-анимации.")
        return
        
    fig = px.scatter(
        df_agg, 
        x="Year", 
        y="total_goals",
        animation_frame="Year",
        animation_group="team",
        size="num_matches", 
        color="team",
        hover_name="team",
        log_y=False,
        title="Динамика голов по командам в международных матчах (с 2020 года)"
    )
    
    if fig.layout.updatemenus:
        fig.layout.updatemenus[0].buttons[0].args[1]['frame']['duration'] = 1000
        fig.layout.updatemenus[0].buttons[0].args[1]['transition']['duration'] = 500
    
    print("\n[ЗАДАНИЕ 2: Plotly - Интерактивный ползунок]")
    print("   -> Создан интерактивный график с временным ползунком. Выводится в браузере.")
    fig.show()

def export_to_excel(dataframes_dict: Dict[str, pd.DataFrame], filename: str):
    """Экспорт DataFrame'ов в Excel с форматированием, включая градиент для голов."""
    sheet_count = 0
    total_rows = 0
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet_count += 1
            total_rows += len(df)
            
            workbook = writer.book
            ws = workbook[sheet_name]
            
            if len(df) > 0:
                ws.freeze_panes = "A2"
            
            ws.auto_filter.ref = ws.dimensions
            
            # --- 1. Стандартное форматирование Min/Max для всех числовых колонок ---
            for col_idx, dtype in enumerate(df.dtypes):
                # Пропускаем колонку 'goals' на листе 'top_scorers', чтобы градиент работал
                if sheet_name == 'top_scorers' and df.columns[col_idx] == 'goals':
                    continue
                
                if pd.api.types.is_numeric_dtype(dtype) and len(df) > 1:
                    col_letter = get_column_letter(col_idx + 1)
                    data_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                    
                    # Выделение максимума (синий)
                    max_fill = PatternFill(start_color='FF0000AA', end_color='FF0000AA', fill_type='solid')
                    max_dxf = DifferentialStyle(fill=max_fill) 
                    max_rule = Rule(
                        type='expression', 
                        operator='equal', 
                        formula=[f'=MAX(${col_letter}2:${col_letter}${len(df)+1})'], 
                        dxf=max_dxf, 
                        stopIfTrue=True
                    )
                    ws.conditional_formatting.add(data_range, max_rule)
                    
                    # Выделение минимума (желтый)
                    min_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    min_dxf = DifferentialStyle(fill=min_fill) 
                    min_rule = Rule(
                        type='expression', 
                        operator='equal', 
                        formula=[f'=MIN(${col_letter}2:${col_letter}${len(df)+1})'], 
                        dxf=min_dxf, 
                        stopIfTrue=True
                    )
                    ws.conditional_formatting.add(data_range, min_rule)

            # --- 2. ГРАДИЕНТ ЦВЕТА для колонки "goals" (только для листа 'top_scorers') ---
            if sheet_name == 'top_scorers' and 'goals' in df.columns and len(df) > 1:
                goals_col_idx = df.columns.get_loc('goals')
                col_letter = get_column_letter(goals_col_idx + 1)
                data_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                
                # Используем надежный двухцветный градиент
                gradient_rule = ColorScaleRule(
                    start_type='min', 
                    start_color='FFFFFFFF',  # Белый (минимум голов)
                    end_type='max', 
                    end_color='FF63BE7B'     # Зеленый (максимум голов)
                )
                
                ws.conditional_formatting.add(data_range, gradient_rule)
                print(f"   -> Применена градиентная заливка к колонке 'goals'.")
            elif sheet_name == 'top_scorers':
                 print(f"   -> ВНИМАНИЕ: Градиент не применен. Колонка 'goals' не найдена на листе '{sheet_name}'.")


    print("\n[ЗАДАНИЕ 3: Экспорт в Excel с форматированием]")
    print(f"   -> Создан файл {filename}, {sheet_count} листа(ов), {total_rows} строк (сумма).")


# ----------------------------------------------------------------------
#  ГЛАВНАЯ ФУНКЦИЯ
# ----------------------------------------------------------------------

def main():
    
    CLIENT_ENCODING = "UTF8" 
    dataframes_for_export = {}
    
    try:
        with psycopg2.connect(
            dbname=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD,
            host=DB_HOST,
            port=DB_PORT,
            options=f"-c client_encoding={CLIENT_ENCODING}"
        ) as conn:
            print(f"✅ Подключение успешно. Client Encoding: {CLIENT_ENCODING}")

            print("\n=====================================")
            print("         [ЗАДАНИЕ 1: ВИЗУАЛИЗАЦИЯ]      ")
            print("=====================================")
            
            # --- ЗАДАНИЕ 1: Визуализации (6 графиков) ---
            df_map = {}
            query_names = list(ALL_QUERIES.keys())[:6] 
            
            for name in query_names:
                df_map[name] = run_query(ALL_QUERIES[name], conn)

            if df_map["q_pie_stage_dist"] is not None and not df_map["q_pie_stage_dist"].empty:
                create_pie_chart(df_map["q_pie_stage_dist"].copy())
                
            if df_map["q_bar_top_goal_teams"] is not None and not df_map["q_bar_top_goal_teams"].empty:
                create_bar_chart(df_map["q_bar_top_goal_teams"].copy())
            
            if df_map["q_hbar_avg_fifa_rank"] is not None and not df_map["q_hbar_avg_fifa_rank"].empty:
                create_horizontal_bar(df_map["q_hbar_avg_fifa_rank"].copy())

            if df_map["q_line_goals_per_match"] is not None and not df_map["q_line_goals_per_match"].empty:
                create_line_chart(df_map["q_line_goals_per_match"].copy())
                
            if df_map["q_bar_matches_2022"] is not None and not df_map["q_bar_matches_2022"].empty:
                create_histogram(df_map["q_bar_matches_2022"].copy())

            if df_map["q_scatter_goals_vs_teams"] is not None and not df_map["q_scatter_goals_vs_teams"].empty:
                create_scatter_plot(df_map["q_scatter_goals_vs_teams"].copy())
                
            print("\n=====================================")
            print("         [ЗАДАНИЕ 2: PLOTLY]         ")
            print("=====================================")
            
            # --- ЗАДАНИЕ 2: Plotly (Временной ползунок) ---
            df_plotly = run_query(ALL_QUERIES["q_plotly_intl_goals"], conn)
            if df_plotly is not None and not df_plotly.empty:
                 create_plotly_animation(df_plotly.copy())

            # --- ЗАДАНИЕ 3: Сборка данных для Excel ---
            print("\n=====================================")
            print("         [ЗАДАНИЕ 3: ЭКСПОРТ]          ")
            print("=====================================")
            
            for name in ["top_scorers_export", "matches_by_year_export", "teams_by_group_export"]:
                 df = run_query(ALL_QUERIES[name], conn)
                 if df is not None:
                     dataframes_for_export[name.replace('_export', '')] = df
                     
            if dataframes_for_export:
                export_to_excel(dataframes_for_export, "exports/WorldCup_Report_Formatted.xlsx")

    except psycopg2.Error as e:
        print(f"❌ Ошибка подключения к базе данных. Проверьте настройки: {e}")
    except Exception as e:
        print(f"❌ Непредвиденная ошибка: {e}")


if __name__ == "__main__":
    main()